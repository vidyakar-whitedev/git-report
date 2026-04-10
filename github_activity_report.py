"""
GitHub Activity Report — Production Grade
==========================================
Collects commits, pull requests, and GitHub Actions workflow runs across one
repository or an entire GitHub organisation, joins the data on commit SHA, and
writes a colour-coded multi-sheet Excel report.

Sheets produced
---------------
  1. Activity        — commits × PRs × workflow runs (joined on commit SHA)
  2. Access Control  — collaborators and teams with permission levels
  3. Failure Summary — failed / timed-out runs with root-cause diagnostics
  4. Failure Alerts  — banner + summary table for all failures found

Environment variables (all resolved at startup)
-----------------------------------------------
Required (at least one):
  GH_PAT          Personal Access Token  (needs repo + workflow + read:org)
                  Falls back to GITHUB_TOKEN when running inside GitHub Actions.

  GH_REPO         Single repo in "owner/repo" format.
                  Omit to scan every repo the token can reach.

  GH_ORG          Organisation login (e.g. "my-company").
                  When set, only repos belonging to this org are processed.
                  ┌─────────────────────────────────────────────────────────┐
                  │  ✏  CHANGE THIS for your organisation                   │
                  │  Set GH_ORG=<your-github-org-name> in your CI secret    │
                  │  or in the workflow env block.                           │
                  └─────────────────────────────────────────────────────────┘

Optional:
  GH_OUTPUT       Output file path          (default: github_activity_report.xlsx)
  GH_MAX_RUNS     Max workflow runs fetched per repo (default: 200)
  GH_LOOKBACK_DAYS  Report runs from last N days (default: 30)
  GH_SINCE        ISO-8601 start date, e.g. "2024-01-01"  (overrides LOOKBACK_DAYS)
  GH_UNTIL        ISO-8601 end date,   e.g. "2024-12-31"
  GH_LOG_LEVEL    Logging verbosity: DEBUG | INFO | WARNING (default: INFO)
  GH_REPORT_TITLE Custom title shown in the Excel banner row
                  ┌─────────────────────────────────────────────────────────┐
                  │  ✏  CHANGE THIS to your organisation / project name     │
                  └─────────────────────────────────────────────────────────┘
"""

from __future__ import annotations

import io
import os
import re
import sys
import time
import logging
import zipfile
from datetime import datetime, timedelta, timezone
import requests
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# ✏  SECTION 1 — ORGANISATION / PROJECT CONFIGURATION
#    These are the values most likely to differ between organisations.
#    Every line marked "✏ CHANGE" should be reviewed before first deployment.
# ══════════════════════════════════════════════════════════════════════════════

# -- Token & target ------------------------------------------------------------

GH_PAT   = os.getenv("GH_PAT") or os.getenv("GITHUB_TOKEN", "")
GH_REPO  = os.getenv("GH_REPO") or os.getenv("GITHUB_REPOSITORY", "")
GH_ORG   = os.getenv("GH_ORG", "")          # ✏ CHANGE: set to your org login

# -- Output & scope ------------------------------------------------------------

OUTPUT        = os.getenv("GH_OUTPUT", "github_activity_report.xlsx")
MAX_RUNS      = int(os.getenv("GH_MAX_RUNS",        "200"))   # ✏ CHANGE: increase for busier orgs
LOOKBACK_DAYS = int(os.getenv("GH_LOOKBACK_DAYS",   "30"))    # ✏ CHANGE: days of history to include
GH_SINCE      = os.getenv("GH_SINCE", "")    # e.g. "2024-01-01" — overrides LOOKBACK_DAYS
GH_UNTIL      = os.getenv("GH_UNTIL", "")    # e.g. "2024-12-31"

# -- Report branding -----------------------------------------------------------

REPORT_TITLE = os.getenv(
    "GH_REPORT_TITLE",
    "GitHub Activity & Workflow Audit Report",   # ✏ CHANGE: your org / project name
)

BASE_URL = "https://api.github.com"

# ══════════════════════════════════════════════════════════════════════════════
# ✏  SECTION 2 — EXCEL COLOUR THEME
#    Change hex colours here to match your organisation's brand palette.
#    Hex codes must be 6-character RRGGBB (no leading #).
# ══════════════════════════════════════════════════════════════════════════════

_BRAND_DARK   = "1F3864"   # ✏ CHANGE: header background       (default: dark navy)
_BRAND_LIGHT  = "EEF2F7"   # ✏ CHANGE: alternate row stripe     (default: light grey-blue)
_SUCCESS_BG   = "C6EFCE"   # ✏ CHANGE: success row background   (default: light green)
_SUCCESS_FG   = "276221"   # ✏ CHANGE: success row text colour  (default: dark green)
_FAILURE_BG   = "FFC7CE"   # ✏ CHANGE: failure row background   (default: light red)
_FAILURE_FG   = "9C0006"   # ✏ CHANGE: failure row text colour  (default: dark red)
_WARN_BG      = "FFEB9C"   # ✏ CHANGE: warning row background   (default: light yellow)
_WARN_FG      = "9C5700"   # ✏ CHANGE: warning row text colour  (default: amber)

_HEADER_FILL  = PatternFill("solid", fgColor=_BRAND_DARK)
_HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
_ALT_FILL     = PatternFill("solid", fgColor=_BRAND_LIGHT)
_SUCCESS_FILL = PatternFill("solid", fgColor=_SUCCESS_BG)
_SUCCESS_FONT = Font(color=_SUCCESS_FG, bold=True, size=10)
_FAILURE_FILL = PatternFill("solid", fgColor=_FAILURE_BG)
_FAILURE_FONT = Font(color=_FAILURE_FG, bold=True, size=10)
_WARN_FILL    = PatternFill("solid", fgColor=_WARN_BG)
_WARN_FONT    = Font(color=_WARN_FG, size=10)
_THIN         = Border(
    left=Side(style="thin",   color="D0D7DE"),
    right=Side(style="thin",  color="D0D7DE"),
    top=Side(style="thin",    color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)

# ══════════════════════════════════════════════════════════════════════════════
# ✏  SECTION 3 — FAILURE FIX-HINT RULES
#    Each entry: ([keyword_list], "human-readable suggestion")
#    Add or edit rules to match the tech stack used in your organisation.
# ══════════════════════════════════════════════════════════════════════════════

_FIX_RULES: list[tuple[list[str], str]] = [
    # ✏ CHANGE: add keywords specific to your build tools / frameworks
    (
        ["install", "pip", "npm", "yarn", "poetry", "dependency",
         "module not found", "importerror", "modulenotfounderror", "no module"],
        "Dependency error — run `pip install -r requirements.txt` or `npm install` "
        "and verify package versions.",
    ),
    (
        ["permission", "forbidden", "401", "403", "unauthorized",
         "access denied", "secret", "token", "credential"],
        "Auth/permission error — verify GH_PAT scopes and that repository "
        "secrets are correctly configured.",
    ),
    (
        ["timeout", "timed out", "timed_out", "deadline exceeded"],
        "Timeout — increase job `timeout-minutes`, optimise long-running steps, "
        "or split the job.",
    ),
    (
        ["syntax", "syntaxerror", "parse error", "unexpected token",
         "invalid syntax", "yaml", "yml"],
        "Syntax error — review workflow YAML and scripts for syntax mistakes.",
    ),
    (
        ["docker", "container", "image", "pull"],
        "Docker/container error — verify image name/tag is correct and accessible.",
    ),
    (
        ["test", "assert", "spec", "jest", "pytest", "unittest", "rspec", "coverage"],
        "Test failure — review failing test output and fix the failing assertions.",
    ),
    (
        ["build", "compile", "tsc", "webpack", "gradle", "maven", "make"],
        "Build error — check compiler output in logs and fix the source.",
    ),
    (
        ["deploy", "release", "publish", "helm", "kubectl", "terraform"],
        "Deployment error — verify credentials, target environment, and release config.",
    ),
]

# ══════════════════════════════════════════════════════════════════════════════
# Logging
# ══════════════════════════════════════════════════════════════════════════════

_LOG_LEVEL = os.getenv("GH_LOG_LEVEL", "INFO").upper()

logging.basicConfig(
    level=getattr(logging, _LOG_LEVEL, logging.INFO),
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════════════════
# HTTP client
# ══════════════════════════════════════════════════════════════════════════════

def _build_session() -> requests.Session:
    """Create an authenticated requests session for the GitHub REST API."""
    session = requests.Session()
    if not GH_PAT:
        log.warning(
            "GH_PAT / GITHUB_TOKEN not set — unauthenticated requests "
            "are rate-limited to 60/hour"
        )
    session.headers.update({
        "Authorization":        f"Bearer {GH_PAT}" if GH_PAT else "",
        "Accept":               "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    })
    return session


SESSION = _build_session()


def _wait_for_rate_limit(headers: dict) -> None:
    """Sleep until the GitHub rate-limit window resets."""
    reset_ts = int(headers.get("X-RateLimit-Reset", time.time() + 60))
    wait_sec = max(reset_ts - int(time.time()), 1)
    log.warning("Rate-limited — sleeping %d s …", wait_sec)
    time.sleep(wait_sec)


def _get(url: str, params: dict | None = None) -> dict | list | None:
    """
    Single GET with automatic rate-limit retry.

    Returns:
        Parsed JSON body, or None on 404.
    Raises:
        RuntimeError: after 5 failed retries.
    """
    for attempt in range(5):
        try:
            resp = SESSION.get(url, params=params or {}, timeout=30)
        except requests.RequestException as exc:
            log.warning("GET %s attempt %d failed: %s", url, attempt + 1, exc)
            time.sleep(2 ** attempt)
            continue
        if resp.status_code == 403 and "rate limit" in resp.text.lower():
            _wait_for_rate_limit(resp.headers)
            continue
        if resp.status_code == 404:
            return None
        resp.raise_for_status()
        return resp.json()
    raise RuntimeError(f"GET {url} failed after 5 retries")


def _paginate(url: str, params: dict | None = None, max_items: int = 0) -> list:
    """
    Follow GitHub's Link-header pagination and return all items.

    Unwraps common envelope keys: workflow_runs, jobs, repositories.
    Returns an empty list on 403 / 404 (no access) rather than raising.
    """
    params = dict(params or {})
    params.setdefault("per_page", 100)
    items:    list      = []
    page_url: str | None = url

    while page_url:
        for attempt in range(5):
            try:
                resp = SESSION.get(page_url, params=params, timeout=30)
            except requests.RequestException as exc:
                log.warning("Paginate %s attempt %d failed: %s", page_url, attempt + 1, exc)
                time.sleep(2 ** attempt)
                continue
            if resp.status_code == 403 and "rate limit" in resp.text.lower():
                _wait_for_rate_limit(resp.headers)
                continue
            if resp.status_code in (403, 404):
                return items
            resp.raise_for_status()
            break

        payload = resp.json()
        if isinstance(payload, dict):
            for key in ("workflow_runs", "jobs", "repositories"):
                if key in payload:
                    payload = payload[key]
                    break
            else:
                payload = []

        items.extend(payload)
        params = {}   # Link: next already encodes query params

        if max_items and len(items) >= max_items:
            return items[:max_items]

        page_url = resp.links.get("next", {}).get("url")

    return items

# ══════════════════════════════════════════════════════════════════════════════
# Repository discovery
# ══════════════════════════════════════════════════════════════════════════════

def fetch_repos() -> list[dict]:
    """
    Return a list of GitHub repository objects to process.

    Resolution order:
      1. GH_REPO  — single specific repo (owner/repo format)
      2. GH_ORG   — all repos in the named organisation
      3. fallback — all repos accessible to the authenticated token

    ✏  CHANGE: if your org uses GitHub Enterprise Server, replace BASE_URL
       at the top of this file with your GHES API endpoint, e.g.:
       https://github.mycompany.com/api/v3
    """
    if GH_REPO:
        if "/" not in GH_REPO:
            log.error(
                "GH_REPO must be 'owner/repo' format.  Got: %r", GH_REPO
            )
            sys.exit(1)
        owner, name = GH_REPO.split("/", 1)
        data = _get(f"{BASE_URL}/repos/{owner}/{name}")
        if data is None:
            log.error("Repository not found or token lacks access: %s", GH_REPO)
            sys.exit(1)
        log.info("Single-repo mode: %s", GH_REPO)
        return [data]

    if GH_ORG:
        log.info("Organisation mode: fetching repos for org '%s' …", GH_ORG)
        repos = _paginate(
            f"{BASE_URL}/orgs/{GH_ORG}/repos",
            params={"type": "all"},   # ✏ CHANGE: use "public" to skip private repos
        )
    else:
        log.info("User mode: fetching all accessible repositories …")
        repos = _paginate(
            f"{BASE_URL}/user/repos",
            params={"affiliation": "owner,collaborator,organization_member"},
        )

    if not repos:
        log.error(
            "No repositories found.  "
            "Check that GH_PAT has the required scopes and GH_ORG / GH_REPO is correct."
        )
        sys.exit(1)

    log.info("  %d repository/repositories found", len(repos))
    return repos


def _repo_meta(repo: dict) -> dict:
    """Extract a normalised metadata dict from a raw GitHub repo object."""
    return {
        "org":            (repo.get("owner") or {}).get("login", ""),
        "name":           repo.get("name", ""),
        "full_name":      repo.get("full_name", ""),
        "visibility":     repo.get("visibility", ""),
        "default_branch": repo.get("default_branch", "main"),
    }

# ══════════════════════════════════════════════════════════════════════════════
# Data fetchers
# ══════════════════════════════════════════════════════════════════════════════

def fetch_branches(org: str, repo: str) -> dict[str, str]:
    """Return {branch_name: head_sha} for every branch in the repository."""
    branches = _paginate(f"{BASE_URL}/repos/{org}/{repo}/branches")
    return {b["name"]: b["commit"]["sha"] for b in branches}


def fetch_commits(org: str, repo: str, branches: dict[str, str]) -> list[dict]:
    """
    Fetch all unique commits across every branch, de-duplicated by SHA.

    A commit reachable from multiple branches is recorded once, attributed
    to the first branch it was encountered on.
    """
    seen: dict[str, dict] = {}
    for branch_name in branches:
        for c in _paginate(
            f"{BASE_URL}/repos/{org}/{repo}/commits",
            params={"sha": branch_name},
        ):
            sha = c["sha"]
            if sha in seen:
                continue
            git_commit = c.get("commit") or {}
            git_author = git_commit.get("author") or {}
            gh_author  = c.get("author") or {}
            seen[sha] = {
                "sha":     sha,
                "author":  gh_author.get("login") or git_author.get("name", ""),
                "message": git_commit.get("message", "").split("\n")[0],
                "date":    git_author.get("date", ""),
                "branch":  branch_name,
            }
    return list(seen.values())


def fetch_pull_requests(org: str, repo: str) -> list[dict]:
    """
    Fetch all pull requests (open, closed, merged).

    Returns the fields required for SHA-based join and Activity columns.
    """
    prs = _paginate(
        f"{BASE_URL}/repos/{org}/{repo}/pulls",
        params={"state": "all", "sort": "updated", "direction": "desc"},
    )
    result = []
    for pr in prs:
        merged_at = pr.get("merged_at") or ""
        result.append({
            "pr_id":        pr.get("number"),
            "pr_title":     pr.get("title", ""),
            "pr_author":    (pr.get("user") or {}).get("login", ""),
            "pr_status":    "merged" if merged_at else pr.get("state", ""),
            "pr_merged":    "Yes" if merged_at else "No",
            "pr_merged_at": merged_at,
            "merge_sha":    pr.get("merge_commit_sha") or "",
            "head_sha":     (pr.get("head") or {}).get("sha", ""),
        })
    return result


def _date_filter_bounds() -> tuple[datetime, datetime | None]:
    """
    Resolve date-filter bounds from environment variables.

    Priority:  GH_SINCE / GH_UNTIL  >  GH_LOOKBACK_DAYS
    Returns:   (since_dt, until_dt) — until_dt is None when no upper bound.
    """
    until_dt: datetime | None = None

    if GH_SINCE:
        since_dt = datetime.fromisoformat(GH_SINCE).replace(tzinfo=timezone.utc)
    else:
        since_dt = datetime.now(timezone.utc) - timedelta(days=LOOKBACK_DAYS)

    if GH_UNTIL:
        until_dt = datetime.fromisoformat(GH_UNTIL).replace(
            hour=23, minute=59, second=59, tzinfo=timezone.utc
        )

    return since_dt, until_dt


def _parse_iso(ts: str) -> datetime:
    """Parse a GitHub API ISO-8601 timestamp (handles trailing Z)."""
    return datetime.fromisoformat(ts.rstrip("Z") + "+00:00")


def _fetch_failure_detail(org: str, repo: str, run_id: int) -> dict:
    """
    Drill into a failed run's jobs/steps and extract a log snippet.

    Log download is best-effort — some token scopes cannot access raw logs.
    Returns a dict with: failed_job, failed_step, log_snippet, error_line,
    suggested_fix.
    """
    jobs = _paginate(f"{BASE_URL}/repos/{org}/{repo}/actions/runs/{run_id}/jobs")

    failed_job  = ""
    failed_step = ""
    log_snippet = ""
    error_line  = "See GitHub Actions logs"

    for job in jobs:
        if job.get("conclusion") in ("failure", "timed_out", "cancelled"):
            failed_job = job.get("name", "unknown job")
            for step in job.get("steps", []):
                if step.get("conclusion") in ("failure", "timed_out"):
                    failed_step = step.get("name", "unknown step")
                    break
            break

    # Attempt to download and parse the log zip
    log_url = f"{BASE_URL}/repos/{org}/{repo}/actions/runs/{run_id}/logs"
    try:
        resp = SESSION.get(log_url, timeout=30, allow_redirects=True)
        if resp.status_code == 200 and resp.content:
            with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
                target = failed_job.lower().replace(" ", "_")
                for fname in zf.namelist():
                    if target in fname.lower() or not failed_job:
                        raw = zf.read(fname).decode("utf-8", errors="replace")
                        # ✏ CHANGE: extend this keyword list with error patterns
                        #   specific to your stack (e.g. "AssertionError", "FATAL")
                        error_lines = [
                            ln.strip() for ln in raw.splitlines()
                            if any(
                                kw in ln.lower()
                                for kw in (
                                    "error", "fatal", "failed", "exception",
                                    "traceback", "stderr", "syntaxerror",
                                )
                            )
                        ]
                        if error_lines:
                            log_snippet = error_lines[0][:200]
                            for ln in error_lines[:10]:
                                m = re.search(r"(?:line |:)(\d+)", ln, re.IGNORECASE)
                                if m:
                                    error_line = f"Line {m.group(1)}"
                                    break
                        break
    except Exception as exc:
        log.debug("Log download skipped for run %d: %s", run_id, exc)

    return {
        "failed_job":    failed_job,
        "failed_step":   failed_step,
        "log_snippet":   log_snippet,
        "error_line":    error_line,
        "suggested_fix": _suggest_fix("failure", failed_job, failed_step, log_snippet),
    }


def _suggest_fix(conclusion: str, job: str, step: str, snippet: str) -> str:
    """Map failure context to a human-readable fix suggestion using _FIX_RULES."""
    if conclusion == "timed_out":
        return "Timeout — increase job `timeout-minutes` or optimise the long-running step."
    if conclusion == "startup_failure":
        return "Workflow startup failure — check YAML syntax and runner availability."
    if conclusion == "cancelled":
        return "Run was cancelled manually or by a newer push — no action needed unless unintended."

    haystack = " ".join([job, step, snippet]).lower()
    for keywords, suggestion in _FIX_RULES:
        if any(kw in haystack for kw in keywords):
            return suggestion

    return "Review the full workflow log in GitHub Actions for the specific error message."


def fetch_workflow_runs(org: str, repo: str) -> list[dict]:
    """
    Fetch workflow runs within the configured date window.

    Pulls up to MAX_RUNS from the API then filters client-side to
    [since_dt, until_dt].  Failed runs are enriched with job/step diagnostics.
    """
    since_dt, until_dt = _date_filter_bounds()
    since_label = since_dt.strftime("%Y-%m-%d")
    until_label = until_dt.strftime("%Y-%m-%d") if until_dt else "now"
    log.info(
        "    Fetching workflow runs (max %d, %s → %s) …",
        MAX_RUNS, since_label, until_label,
    )

    raw_runs = _paginate(
        f"{BASE_URL}/repos/{org}/{repo}/actions/runs",
        max_items=MAX_RUNS,
    )

    # Filter to the requested date window
    filtered = []
    for run in raw_runs:
        ts_str = run.get("created_at") or run.get("run_started_at") or ""
        if not ts_str:
            continue
        run_dt = _parse_iso(ts_str)
        if run_dt < since_dt:
            continue
        if until_dt and run_dt > until_dt:
            continue
        filtered.append(run)

    log.info("      %d run(s) in window", len(filtered))

    result: list[dict] = []
    for run in filtered:
        conclusion = run.get("conclusion") or ""
        run_id     = run["id"]

        if conclusion in ("failure", "timed_out", "startup_failure"):
            log.info("      fetching failure detail for run %d …", run_id)
            detail = _fetch_failure_detail(org, repo, run_id)
        else:
            detail = {
                "failed_job":    "",
                "failed_step":   "",
                "log_snippet":   "",
                "error_line":    "",
                "suggested_fix": "",
            }

        # Build a readable failure reason string
        if conclusion == "failure":
            parts = []
            if detail["failed_job"]:
                parts.append(f"Job '{detail['failed_job']}'")
            if detail["failed_step"]:
                parts.append(f"step '{detail['failed_step']}' failed")
            if detail["log_snippet"]:
                parts.append(f"— {detail['log_snippet'][:120]}")
            failure_reason = " ".join(parts) if parts else "Failure (see logs)"
            detail["suggested_fix"] = _suggest_fix(
                "failure", detail["failed_job"], detail["failed_step"], detail["log_snippet"]
            )
        elif conclusion == "timed_out":
            failure_reason = (
                f"Timed out in job '{detail['failed_job']}'"
                if detail["failed_job"] else "Timed out"
            )
            detail["suggested_fix"] = _suggest_fix("timed_out", detail["failed_job"], "", "")
        elif conclusion == "startup_failure":
            failure_reason = "Workflow startup failure"
            detail["suggested_fix"] = _suggest_fix("startup_failure", "", "", "")
        elif conclusion == "cancelled":
            failure_reason = "Cancelled"
            detail["suggested_fix"] = _suggest_fix("cancelled", "", "", "")
        else:
            failure_reason = ""

        if failure_reason:
            log.warning(
                "WORKFLOW FAILURE in %s/%s  (Run ID: %d)  — %s",
                org, repo, run_id, failure_reason,
            )

        result.append({
            "run_id":         run_id,
            "workflow":       run.get("name", ""),
            "status":         run.get("status", ""),
            "conclusion":     conclusion,
            "event":          run.get("event", ""),
            "head_sha":       run.get("head_sha", ""),
            "run_started_at": run.get("run_started_at", ""),
            "run_author":     (
                run.get("triggering_actor") or run.get("actor") or {}
            ).get("login", ""),
            "failure_reason": failure_reason,
            **detail,
        })

    return result


def fetch_access_control(org: str, repo: str) -> list[dict]:
    """
    Fetch collaborators (individual users) and teams with permission levels.

    Returns an empty list for personal repos — team endpoints return 404
    which _paginate handles gracefully.

    ✏  CHANGE: If your org uses SAML SSO, the collaborator list only includes
       members who have authorised the token — this is a GitHub API limitation.
    """
    records: list[dict] = []
    _PERM_MAP = {
        "admin": "Admin", "maintain": "Maintain",
        "push": "Write", "triage": "Triage", "pull": "Read",
    }

    for user in _paginate(
        f"{BASE_URL}/repos/{org}/{repo}/collaborators",
        params={"affiliation": "all"},
    ):
        perms = user.get("permissions") or {}
        if perms.get("admin"):
            level = "Admin"
        elif perms.get("maintain"):
            level = "Maintain"
        elif perms.get("push"):
            level = "Write"
        elif perms.get("triage"):
            level = "Triage"
        else:
            level = "Read"
        records.append({
            "entity_name": user.get("login", ""),
            "entity_type": "User",
            "permission":  level,
            "has_admin":   "Yes" if perms.get("admin") else "No",
            "can_delete":  "Yes" if perms.get("admin") else "No",
        })

    for team in _paginate(f"{BASE_URL}/repos/{org}/{repo}/teams"):
        perm  = team.get("permission", "pull")
        level = _PERM_MAP.get(perm, perm.capitalize())
        records.append({
            "entity_name": team.get("name", ""),
            "entity_type": "Team",
            "permission":  level,
            "has_admin":   "Yes" if perm == "admin" else "No",
            "can_delete":  "Yes" if perm == "admin" else "No",
        })

    return records

# ══════════════════════════════════════════════════════════════════════════════
# ✏  SECTION 4 — COLUMN DEFINITIONS
#    Add, remove, or rename columns here.
#    Column names here must exactly match the keys used in the row-builder
#    functions below.
# ══════════════════════════════════════════════════════════════════════════════

ACTIVITY_COLUMNS = [
    # Repository metadata
    "Repository", "Organization", "Visibility", "Default Branch",
    # Commit
    "Commit ID", "Commit Message", "Author", "Date", "Branch",
    # Pull Request
    "PR ID", "PR Title", "PR Author", "PR Status", "PR Merged", "PR Merge Date",
    # Workflow run
    "Workflow Name", "Workflow Run ID", "Trigger Event",
    "Run Started At", "Workflow Status", "Workflow Conclusion",
    # Failure diagnostics
    "Failure Reason", "Failed Job", "Failed Step", "Error Line", "Suggested Fix",
]

ACCESS_COLUMNS = [
    "Repository", "Organization",
    "User/Team Name", "Type", "Permission Level", "Has Admin", "Has Delete Access",
]

FAILURE_COLUMNS = [
    "Repository", "Organization",
    "Workflow Name", "Workflow Run ID", "Trigger Event",
    "Run Started At", "Workflow Status", "Workflow Conclusion",
    "Failure Reason", "Failed Job", "Failed Step", "Error Line", "Suggested Fix",
]

ALERTS_COLUMNS = [
    "#", "Repository", "Organization", "Workflow Name", "Run ID",
    "Trigger", "Run Started At", "Conclusion",
    "Failure Reason", "Failed Job", "Failed Step", "Error Line", "Suggested Fix",
]

# ══════════════════════════════════════════════════════════════════════════════
# Dataset assembly — join on commit SHA
# ══════════════════════════════════════════════════════════════════════════════

def build_activity_rows(
    meta:    dict,
    commits: list[dict],
    prs:     list[dict],
    runs:    list[dict],
) -> list[dict]:
    """
    Produce the Activity sheet rows by joining commits, PRs and workflow runs
    on commit SHA.

    Join strategy:
      1. Commits are the primary anchor.  Each commit is expanded by its
         matched PRs × matched workflow runs (Cartesian product per commit).
      2. PRs whose SHA doesn't match any fetched commit appear as orphan rows.
      3. Workflow runs whose head_sha doesn't match any commit appear as orphan rows.
    """
    org  = meta["org"]
    repo = meta["name"]
    vis  = meta["visibility"]
    defb = meta["default_branch"]

    # Build SHA lookup indexes
    sha_to_prs: dict[str, list[dict]] = {}
    for pr in prs:
        for key in ("merge_sha", "head_sha"):
            sha = pr.get(key, "")
            if sha:
                sha_to_prs.setdefault(sha, []).append(pr)

    sha_to_runs: dict[str, list[dict]] = {}
    for run in runs:
        sha = run.get("head_sha", "")
        if sha:
            sha_to_runs.setdefault(sha, []).append(run)

    covered_prs:  set[int] = set()
    covered_runs: set[int] = set()
    rows: list[dict] = []

    def _pr_cols(pr: dict | None) -> dict:
        if not pr:
            return {
                "PR ID": "", "PR Title": "", "PR Author": "",
                "PR Status": "", "PR Merged": "", "PR Merge Date": "",
            }
        return {
            "PR ID":        pr["pr_id"],
            "PR Title":     pr["pr_title"],
            "PR Author":    pr["pr_author"],
            "PR Status":    pr["pr_status"],
            "PR Merged":    pr["pr_merged"],
            "PR Merge Date": pr["pr_merged_at"],
        }

    def _run_cols(run: dict | None) -> dict:
        if not run:
            return {
                "Workflow Name": "",    "Workflow Run ID": "",
                "Trigger Event": "",    "Run Started At": "",
                "Workflow Status": "",  "Workflow Conclusion": "",
                "Failure Reason": "",   "Failed Job": "",
                "Failed Step": "",      "Error Line": "",
                "Suggested Fix": "",
            }
        return {
            "Workflow Name":       run["workflow"],
            "Workflow Run ID":     run["run_id"],
            "Trigger Event":       run["event"],
            "Run Started At":      run.get("run_started_at", ""),
            "Workflow Status":     run["status"],
            "Workflow Conclusion": run["conclusion"],
            "Failure Reason":      run.get("failure_reason", ""),
            "Failed Job":          run.get("failed_job", ""),
            "Failed Step":         run.get("failed_step", ""),
            "Error Line":          run.get("error_line", ""),
            "Suggested Fix":       run.get("suggested_fix", ""),
        }

    # 1. Commit-anchored rows
    for commit in commits:
        sha          = commit["sha"]
        matched_prs  = sha_to_prs.get(sha)  or [None]
        matched_runs = sha_to_runs.get(sha) or [None]

        base = {
            "Repository":     repo,
            "Organization":   org,
            "Visibility":     vis,
            "Default Branch": defb,
            "Commit ID":      sha,
            "Commit Message": commit["message"],
            "Author":         commit["author"],
            "Date":           commit["date"],
            "Branch":         commit["branch"],
        }

        for pr in matched_prs:
            if pr:
                covered_prs.add(pr["pr_id"])
            for run in matched_runs:
                if run:
                    covered_runs.add(run["run_id"])
                rows.append({**base, **_pr_cols(pr), **_run_cols(run)})

    # 2. Orphan PRs (no matching commit in window)
    for pr in prs:
        if pr["pr_id"] not in covered_prs:
            rows.append({
                "Repository": repo,   "Organization": org,
                "Visibility": vis,    "Default Branch": defb,
                "Commit ID": "",      "Commit Message": "",
                "Author": "",         "Date": "",         "Branch": "",
                **_pr_cols(pr),       **_run_cols(None),
            })

    # 3. Orphan workflow runs (no matching commit in window)
    for run in runs:
        if run["run_id"] not in covered_runs:
            rows.append({
                "Repository": repo,   "Organization": org,
                "Visibility": vis,    "Default Branch": defb,
                "Commit ID":  run["head_sha"],
                "Commit Message": "", "Author": run["run_author"],
                "Date": "",           "Branch": "",
                **_pr_cols(None),     **_run_cols(run),
            })

    return rows


def build_access_rows(org: str, repo: str, access: list[dict]) -> list[dict]:
    """Flatten access-control records into ACCESS_COLUMNS shape."""
    return [
        {
            "Repository":        repo,
            "Organization":      org,
            "User/Team Name":    a["entity_name"],
            "Type":              a["entity_type"],
            "Permission Level":  a["permission"],
            "Has Admin":         a["has_admin"],
            "Has Delete Access": a["can_delete"],
        }
        for a in access
    ]


def build_failure_rows(org: str, repo: str, runs: list[dict]) -> list[dict]:
    """Extract failed / timed-out runs into FAILURE_COLUMNS shape."""
    return [
        {
            "Repository":          repo,
            "Organization":        org,
            "Workflow Name":       run["workflow"],
            "Workflow Run ID":     run["run_id"],
            "Trigger Event":       run["event"],
            "Run Started At":      run.get("run_started_at", ""),
            "Workflow Status":     run["status"],
            "Workflow Conclusion": run["conclusion"],
            "Failure Reason":      run.get("failure_reason", ""),
            "Failed Job":          run.get("failed_job", ""),
            "Failed Step":         run.get("failed_step", ""),
            "Error Line":          run.get("error_line", ""),
            "Suggested Fix":       run.get("suggested_fix", ""),
        }
        for run in runs
        if run["conclusion"] in ("failure", "timed_out", "startup_failure")
    ]

# ══════════════════════════════════════════════════════════════════════════════
# Excel writer
# ══════════════════════════════════════════════════════════════════════════════

def _sort_rows(rows: list[dict], date_col: str) -> list[dict]:
    """Sort rows latest-first by *date_col* (ISO-8601).  Empty dates go last."""
    return sorted(rows, key=lambda r: r.get(date_col) or "0000", reverse=True)


def _write_sheet(
    ws,
    columns:        list[str],
    rows:           list[dict],
    conclusion_col: str = "",
    date_col:       str = "",
) -> None:
    """
    Write a data sheet sorted latest-first with full-row colour coding.

    Every cell in a row receives the same background fill AND the matching
    coloured font so the entire row reads as green (success), red (failure),
    or yellow (cancelled/skipped).  The conclusion cell is additionally bold.

    Row colours:
        failure / timed_out / startup_failure  →  red
        success                                →  green
        cancelled / skipped                    →  yellow
        no conclusion (even rows)              →  light grey-blue stripe
    """
    if date_col:
        rows = _sort_rows(rows, date_col)

    # Header
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _THIN
    ws.row_dimensions[1].height = 30

    col_widths = [len(c) + 2 for c in columns]

    # Data rows
    for row_idx, row_data in enumerate(rows, start=2):
        conclusion = str(row_data.get(conclusion_col, "")).lower() if conclusion_col else ""
        is_alt     = (row_idx % 2 == 0)

        # Resolve fill + font for the whole row from the conclusion value
        if conclusion in ("failure", "timed_out", "startup_failure"):
            row_fill = _FAILURE_FILL
            row_font = _FAILURE_FONT
        elif conclusion == "success":
            row_fill = _SUCCESS_FILL
            row_font = _SUCCESS_FONT
        elif conclusion in ("cancelled", "skipped"):
            row_fill = _WARN_FILL
            row_font = _WARN_FONT
        else:
            row_fill = _ALT_FILL if is_alt else None
            row_font = None

        for col_idx, col_name in enumerate(columns, start=1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = _THIN

            if row_fill:
                cell.fill = row_fill

            if row_font:
                # Bold only on the conclusion cell; regular weight everywhere else
                if col_name == conclusion_col:
                    cell.font = row_font
                else:
                    cell.font = Font(color=row_font.color.rgb, size=10)

            col_widths[col_idx - 1] = max(col_widths[col_idx - 1], min(len(str(val)), 60))

    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 62)

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions


def _write_alerts_sheet(ws, alert_rows: list[dict]) -> None:
    """
    Write the Failure Alerts sheet — latest failures at the top.

    Layout:
        Row 1  — merged banner: "⚠  N FAILURE(S)" or "✔  NO FAILURES"
        Row 2  — generated timestamp (right-aligned)
        Row 3  — column headers
        Row 4+ — one alert per row, sorted latest-first, all red
    """
    # Sort latest run first so the most recent failure is always at the top
    alert_rows = _sort_rows(alert_rows, "Run Started At")
    # Re-number the # column after sort
    for i, row in enumerate(alert_rows, start=1):
        row["#"] = i

    n_cols = len(ALERTS_COLUMNS)
    total  = len(alert_rows)

    banner_text = (
        f"⚠   {total} WORKFLOW FAILURE(S) DETECTED"
        if total else "✔   NO WORKFLOW FAILURES DETECTED"
    )
    banner_fill = (
        PatternFill("solid", fgColor="9C0006")
        if total else PatternFill("solid", fgColor="375623")
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    banner = ws.cell(row=1, column=1, value=banner_text)
    banner.font      = Font(color="FFFFFF", bold=True, size=14)
    banner.fill      = banner_fill
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    now_utc = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ts = ws.cell(row=2, column=1, value=f"Generated: {now_utc}  |  {REPORT_TITLE}")
    ts.font      = Font(color="9C0006" if total else "375623", italic=True, size=9)
    ts.alignment = Alignment(horizontal="right")
    ws.row_dimensions[2].height = 14

    HDR_FILL = PatternFill("solid", fgColor="7B0000")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    for col_idx, col_name in enumerate(ALERTS_COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _THIN
    ws.row_dimensions[3].height = 24

    ROW_FILL = PatternFill("solid", fgColor="FFC7CE")
    ALT_FILL = PatternFill("solid", fgColor="FFD7DC")
    ROW_FONT = Font(color="9C0006", size=10)
    col_widths = [len(c) + 2 for c in ALERTS_COLUMNS]

    for row_idx, row_data in enumerate(alert_rows, start=4):
        fill = ALT_FILL if row_idx % 2 == 0 else ROW_FILL
        for col_idx, col_name in enumerate(ALERTS_COLUMNS, start=1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.font      = ROW_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = _THIN
            col_widths[col_idx - 1] = max(col_widths[col_idx - 1], min(len(str(val)), 60))

    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 62)

    ws.freeze_panes = "A4"
    if alert_rows:
        ws.auto_filter.ref = (
            f"A3:{get_column_letter(n_cols)}{3 + len(alert_rows)}"
        )


def _write_cover_sheet(ws) -> None:
    """
    Write a Cover / Summary sheet as the first visible tab.

    ✏  CHANGE: update the contact details and logo colour below.
    """
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 50

    now_utc  = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    since_dt, until_dt = _date_filter_bounds()
    date_range = (
        f"{since_dt.strftime('%Y-%m-%d')} → "
        f"{until_dt.strftime('%Y-%m-%d') if until_dt else 'now'}"
    )

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.merge_cells("B2:C2")
    title_cell = ws.cell(row=2, column=2, value=REPORT_TITLE)
    title_cell.font      = Font(color="FFFFFF", bold=True, size=18)
    title_cell.fill      = PatternFill("solid", fgColor=_BRAND_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 44

    # ── Info block ────────────────────────────────────────────────────────────
    info_rows = [
        ("Generated",    now_utc),
        ("Date Range",   date_range),
        ("Organisation", GH_ORG or GH_REPO or "All accessible repos"),
        ("Max Runs",     str(MAX_RUNS)),
        # ✏ CHANGE: replace with your team's contact details
        ("Maintained by", "DevOps / Platform Engineering team"),
        ("Contact",       "devops@yourcompany.com"),
    ]

    for i, (label, value) in enumerate(info_rows, start=4):
        label_cell = ws.cell(row=i, column=2, value=label)
        value_cell = ws.cell(row=i, column=3, value=value)
        label_cell.font      = Font(bold=True, size=10, color=_BRAND_DARK)
        value_cell.font      = Font(size=10)
        label_cell.alignment = Alignment(vertical="center")
        value_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[i].height = 18

    # ── Sheet legend ──────────────────────────────────────────────────────────
    legend_title = ws.cell(row=11, column=2, value="Sheets in this workbook")
    legend_title.font      = Font(bold=True, size=11, color=_BRAND_DARK)
    legend_title.alignment = Alignment(vertical="center")
    ws.row_dimensions[11].height = 20

    legends = [
        ("Activity",        "All commits, PRs, and workflow runs — latest first"),
        ("Access Control",  "Collaborators and teams with permission levels"),
        ("Failure Summary", "All failed/timed-out runs with root-cause diagnostics"),
        ("Failure Alerts",  "Banner + alert table for all failures detected"),
    ]
    for i, (sheet, desc) in enumerate(legends, start=12):
        sheet_cell = ws.cell(row=i, column=2, value=sheet)
        desc_cell  = ws.cell(row=i, column=3, value=desc)
        sheet_cell.font      = Font(bold=True, size=10)
        desc_cell.font       = Font(size=10)
        sheet_cell.alignment = Alignment(vertical="center")
        desc_cell.alignment  = Alignment(vertical="center")
        ws.row_dimensions[i].height = 18

    # ── Colour key ────────────────────────────────────────────────────────────
    key_title = ws.cell(row=17, column=2, value="Row colour key")
    key_title.font      = Font(bold=True, size=11, color=_BRAND_DARK)
    key_title.alignment = Alignment(vertical="center")
    ws.row_dimensions[17].height = 20

    colour_key = [
        (_SUCCESS_BG, _SUCCESS_FG, "success   — workflow run completed successfully"),
        (_FAILURE_BG, _FAILURE_FG, "failure   — workflow run failed or timed out"),
        (_WARN_BG,    _WARN_FG,    "cancelled — workflow run was cancelled or skipped"),
        (_BRAND_LIGHT, _BRAND_DARK, "no status — commit / PR row with no associated run"),
    ]
    for i, (bg, fg, label) in enumerate(colour_key, start=18):
        swatch = ws.cell(row=i, column=2, value="")
        swatch.fill        = PatternFill("solid", fgColor=bg)
        swatch.border      = _THIN
        desc_cell          = ws.cell(row=i, column=3, value=label)
        desc_cell.font     = Font(color=fg, size=10)
        ws.row_dimensions[i].height = 18


def save_excel(
    activity_rows: list[dict],
    access_rows:   list[dict],
    failure_rows:  list[dict],
    alert_rows:    list[dict],
    path:          str,
) -> None:
    """
    Write all sheets to an Excel workbook at *path*.

    Sheet order:  Cover → Activity → Access Control → Failure Summary → Failure Alerts
    """
    log.info("Writing Excel report → %s", path)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, cols in [
            ("Cover",          []),
            ("Activity",       ACTIVITY_COLUMNS),
            ("Access Control", ACCESS_COLUMNS),
            ("Failure Summary", FAILURE_COLUMNS),
            ("Failure Alerts", ALERTS_COLUMNS),
        ]:
            pd.DataFrame(columns=cols).to_excel(
                writer, index=False, sheet_name=sheet_name
            )

        wb = writer.book

        _write_cover_sheet(wb["Cover"])

        _write_sheet(
            wb["Activity"], ACTIVITY_COLUMNS, activity_rows,
            conclusion_col="Workflow Conclusion",
            date_col="Date",
        )
        _write_sheet(wb["Access Control"], ACCESS_COLUMNS, access_rows)
        _write_sheet(
            wb["Failure Summary"], FAILURE_COLUMNS, failure_rows,
            conclusion_col="Workflow Conclusion",
            date_col="Run Started At",
        )
        _write_alerts_sheet(wb["Failure Alerts"], alert_rows)

        # ✏  CHANGE: tab colours to match your brand
        wb["Cover"].sheet_properties.tabColor          = _BRAND_DARK
        wb["Activity"].sheet_properties.tabColor       = "1F6B9E"   # steel blue
        wb["Access Control"].sheet_properties.tabColor = "375623"   # dark green
        wb["Failure Summary"].sheet_properties.tabColor = "9C0006"  # dark red
        wb["Failure Alerts"].sheet_properties.tabColor  = "FF0000"  # bright red

        wb.active = wb["Failure Alerts"] if alert_rows else wb["Activity"]

    log.info(
        "Report saved → %s  |  Activity: %d  |  Access: %d  |  Failures: %d  |  Alerts: %d",
        path, len(activity_rows), len(access_rows), len(failure_rows), len(alert_rows),
    )

# ══════════════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    log.info("=== %s ===", REPORT_TITLE)
    log.info("Output : %s", OUTPUT)

    if not GH_PAT:
        log.warning(
            "GH_PAT / GITHUB_TOKEN not set — unauthenticated requests are "
            "rate-limited to 60/hour.  Set GH_PAT for production use."
        )

    repos = fetch_repos()

    all_activity: list[dict] = []
    all_access:   list[dict] = []
    all_failures: list[dict] = []
    all_alerts:   list[dict] = []

    for repo in repos:
        meta = _repo_meta(repo)
        org  = meta["org"]
        name = meta["name"]

        log.info("--- %s/%s  [%s] ---", org, name, meta["visibility"])

        log.info("  Fetching branches …")
        branches = fetch_branches(org, name)
        log.info("    %d branch(es)", len(branches))

        log.info("  Fetching commits …")
        commits = fetch_commits(org, name, branches)
        log.info("    %d unique commit(s)", len(commits))

        log.info("  Fetching pull requests …")
        prs = fetch_pull_requests(org, name)
        log.info("    %d PR(s)", len(prs))

        runs = fetch_workflow_runs(org, name)

        log.info("  Fetching access control …")
        access = fetch_access_control(org, name)
        log.info("    %d user(s)/team(s)", len(access))

        all_activity.extend(build_activity_rows(meta, commits, prs, runs))
        all_access.extend(build_access_rows(org, name, access))
        all_failures.extend(build_failure_rows(org, name, runs))

        base_idx = len(all_alerts) + 1
        for idx, run in enumerate(
            (r for r in runs if r["conclusion"] in ("failure", "timed_out", "startup_failure")),
            start=base_idx,
        ):
            all_alerts.append({
                "#":              idx,
                "Repository":     name,
                "Organization":   org,
                "Workflow Name":  run["workflow"],
                "Run ID":         run["run_id"],
                "Trigger":        run["event"],
                "Run Started At": run.get("run_started_at", ""),
                "Conclusion":     run["conclusion"],
                "Failure Reason": run.get("failure_reason", run["conclusion"]),
                "Failed Job":     run.get("failed_job", ""),
                "Failed Step":    run.get("failed_step", ""),
                "Error Line":     run.get("error_line", ""),
                "Suggested Fix":  run.get("suggested_fix", ""),
            })

    # Summary log
    log.info("=== Summary ===")
    log.info("  Repositories : %d", len(repos))
    log.info("  Activity rows: %d", len(all_activity))
    log.info("  Access rows  : %d", len(all_access))
    log.info("  Failure rows : %d", len(all_failures))
    log.info("  Alert rows   : %d", len(all_alerts))

    if all_alerts:
        log.warning("=== FAILURE ALERTS ===")
        for a in all_alerts:
            log.warning(
                "  [%d] %s/%s | Run ID: %s | %s | %s",
                a["#"], a["Organization"], a["Repository"],
                a["Run ID"], a["Workflow Name"], a["Failure Reason"],
            )

    save_excel(all_activity, all_access, all_failures, all_alerts, OUTPUT)

    print(f"\nDone. Report written to: {OUTPUT}")

    if all_alerts:
        sep = "=" * 62
        print(f"\n{sep}")
        print(f"  ⚠  {len(all_alerts)} WORKFLOW FAILURE(S) DETECTED")
        print(sep)
        for a in all_alerts:
            print(
                f"  [{a['#']:>2}] {a['Organization']}/{a['Repository']}"
                f"  |  Run {a['Run ID']}"
                f"  |  {a['Workflow Name']}"
                f"  |  {a['Failure Reason']}"
            )
        print(f"\n  See the 'Failure Alerts' sheet in {OUTPUT} for full details.")


if __name__ == "__main__":
    main()
